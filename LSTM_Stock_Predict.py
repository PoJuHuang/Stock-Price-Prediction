import numpy as np
import pandas as pd
from sklearn.preprocessing import MinMaxScaler
from sklearn.model_selection import train_test_split
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import LSTM, Dense
from tensorflow.keras.metrics import MeanSquaredError, MeanAbsoluteError
from tensorflow.keras.models import load_model
import tensorflow as tf
import matplotlib.pyplot as plt
import os
import matplotlib.dates as mdates
from datetime import datetime
from enum import Enum
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# [Paper1]: 2020 Stock Price Prediction Based on LSTM Neural Network: the Effectiveness of News Sentiment Analysis
# [Paper2]: 2021 Stock Price Prediction Using Sentiment Analysis

# 定義一個枚舉類
class Stock(Enum):
    Tesla = 1
    Microsoft = 2
    Amazon = 3
class Model(Enum):
    Training = 1
    Predict = 2

# =========自行設定參數=============
# 1.[自行更改]設定讀取股票名稱，自行更改
# TARGET = Stock.Tesla
# TARGET = Stock.Microsoft
TARGET = Stock.Amazon

# 2.[自行更改] 要訓練模型還是預測，訓練請用2015-2020期間股票，預測可用2023.07-2023.12期間股票
ModelType = Model.Training
# ModelType = Model.Predict

# 3.[自行更改] 設定讀取股票期間
# [Paper1]使用'2015-2020' or '2023.07-2023.12'
# [Paper2]使用2020_AMZN_5min
Stock_Period = '2015-2020'
# Stock_Period = '2023.07-2023.12'
# Stock_Period = '2020_AMZN_5min'

# 5.[自行更改]設定預測價格方式 1.Close_Price 2.Return_Price
PRICE_Method = 'Close_Price'
# PRICE_Method = 'Return_Price'

# 6.[自行更改]LSTM參數設定，第二篇Paper使用的Time_Step是7
LSTM_Params = {
    'Epochs': 400,
    'Batch_Size': 256,
    'Time_Step': 7
}
# ====================================
# 4.指定讀取模型文件路徑，已經設定好。
model_file_with = ''
model_file_without = ''

# 預測Close
if ((PRICE_Method == 'Close_Price') and (ModelType == Model.Predict)):
    if ((TARGET == Stock.Microsoft)):
        # 指定讀取模型文件路徑，有情感分數模型
        model_file_with = './Models/Predict_Close/6.2024-03-23_09-09-25_Microsoft_with_model.h5'  
        # 指定讀取模型文件路徑，無情感分數模型
        model_file_without = './Models/Predict_Close/6.2024-03-23_09-09-25_Microsoft_without_model.h5' 
    elif ((TARGET == Stock.Tesla)):
        # 指定讀取模型文件路徑，有情感分數模型
        model_file_with = './Models/Predict_Close/6.2024-03-24_20-24-36_Stock.Tesla_with_model.h5'  
        # 指定讀取模型文件路徑，無情感分數模型
        model_file_without = './Models/Predict_Close/6.2024-03-24_20-24-36_Stock.Tesla_without_model.h5' 
    elif ((TARGET == Stock.Amazon)):
        # 指定讀取模型文件路徑，有情感分數模型
        model_file_with = './Models/Predict_Close/6.2024-03-24_15-05-05_Stock.Amazon_with_model.h5'  
        # 指定讀取模型文件路徑，無情感分數模型
        model_file_without = './Models/Predict_Close/6.2024-03-24_15-05-05_Stock.Amazon_without_model.h5'  
# 預測Return
if ((PRICE_Method == 'Return_Price') and (ModelType == Model.Predict)):
    if ((TARGET == Stock.Microsoft)):
        # 指定讀取模型文件路徑，有情感分數模型
        model_file_with = './Models/Predict_Return/6.2024-03-24_22-32-13_Stock.Microsoft_with_model.h5'  
        # 指定讀取模型文件路徑，無情感分數模型
        model_file_without = './Models/Predict_Return/6.2024-03-24_22-32-13_Stock.Microsoft_without_model.h5' 
    elif ((TARGET == Stock.Tesla)):
        # 指定讀取模型文件路徑，有情感分數模型
        model_file_with = './Models/Predict_Return/6.2024-03-24_22-29-53_Stock.Tesla_with_model.h5'  
        # 指定讀取模型文件路徑，無情感分數模型
        model_file_without = './Models/Predict_Return/6.2024-03-24_22-29-53_Stock.Tesla_without_model.h5' 
    elif ((TARGET == Stock.Amazon)):
        # 指定讀取模型文件路徑，有情感分數模型
        model_file_with = './Models/Predict_Return/6.2024-03-24_22-33-23_Stock.Amazon_with_model.h5'  
        # 指定讀取模型文件路徑，無情感分數模型
        model_file_without = './Models/Predict_Return/6.2024-03-24_22-33-23_Stock.Amazon_without_model.h5'  
# ====================================
file_path = ''
# [自行更改]指定讀取CSV檔案路徑，已經設定好。

if (TARGET == Stock.Tesla):   
    file_path = f'./Stock_Datas/{Stock_Period}/TSLA_with_Sentiment.csv'
elif (TARGET == Stock.Microsoft):
    file_path = f'./Stock_Datas/{Stock_Period}/MSFT_with_Sentiment.csv'
elif (TARGET == Stock.Amazon):
    # 第二篇Paper資料，使用AMZN 五分鐘股價資料
    if (Stock_Period == '2020_AMZN_5min'):
        file_path = f'./Stock_Datas/{Stock_Period}/2020_AMZN_5min_with_Sentiment_Add_GoogleNews.csv'
    else:
        file_path = f'./Stock_Datas/{Stock_Period}/AMZN_with_Sentiment.csv'
else:
    print ('讀取路徑異常!')

print (f'讀取路徑: {file_path}')
# ====================================

# 建立目標文件夾路徑
# 獲取當前日期和時間
now = datetime.now()
formatted_now = now.strftime("%Y-%m-%d_%H-%M-%S")

# 建立目標文件夾路徑
save_path = f'./Output_LSTM_Results/{formatted_now}_{TARGET}_{ModelType}_{PRICE_Method}/'
os.makedirs(save_path, exist_ok=True)  # 如果文件夾不存在則創建

# 1. Epochs代表整個訓練數據集被重復運算的次數。在這裡設置為60，意味著整個數據集將被用來訓練模型60次。
# 每次Epoch結束，模型學習到的參數將用來在下一次的Epoch中進行更新和改進。多個Epochs有助於模型更好地學習數據特徵，但過多則可能導致過擬合。

# 2. Batch Size指的是在更新模型參數前用於計算梯度的訓練樣本數量。在這裡設置為256，意味著每次向前和向後傳播將使用256個訓練樣本來更新權重。
# 較小的批次大小通常可以提供更穩定和精確的權重更新，但訓練時間會增長；較大的批次大小可以加快訓練速度，但可能會導致最終模型的性能下降。

# 3. Time Step是指LSTM網絡在處理序列數據時一次考慮的時間點數量。在這裡設置為2，表示模型在每個時間點只考慮兩個時間步長的輸入數據。在處理時間序列數據，如股票價格或語言文本時，Time Step決定了模型如何從序列中學習信息。
# 較小的時間步長意味著模型每次只會看到序列中的一小段信息，而較大的時間步長允許模型一次考慮更長的序列部分。
# 利用n天資料，預測第n+1天，每個序列的時間步長，n = time_step
# 定義LSTM參數

def MAIN(file_path):
    '''主程式'''

    # 文件保存路徑
    file_name = f'{save_path}/0.{TARGET}_LSTM_parameters.txt'
    # 將參數寫入TXT文件
    with open(file_name, 'w') as file:
        for param, value in LSTM_Params.items():
            file.write(f'{param} = {value}\n')
    # 設置隨機種子以確保可重現性
    np.random.seed(42)

    # 生成模擬股價數據和新聞情感分數
    # days = 1000  # 模擬數據的天數
    # stock_prices = np.sin(np.arange(days) * 0.02) + np.random.normal(0, 0.1, days)  # 模擬的股價數據
    # news_sentiment = np.random.uniform(-1, 1, days)  # 模擬的新聞情感分數
    # ======LSTM訓練===============================================

    # 1. 讀取資料
    # 使用pandas讀取CSV檔案
    data = pd.read_csv(file_path)
    # 儲存變數到Txt檔
    # saveToTxt(data,'data.txt')

    # 選取 'Open', 'High', 'Low', 'Close', 和 'Volume' 欄位
    selected_data = data[['Open', 'High', 'Low', 'Close', 'Volume']]

    # Tesla 股票分割，還原原始股價
    if (TARGET == Stock.Tesla):
        # 將 'Open', 'High', 'Low', 'Close' 四个欄位的數據都乘以15
        selected_data[['Open', 'High', 'Low', 'Close']] = selected_data[['Open', 'High', 'Low', 'Close']] * 15
    # Amazon 股票分割，還原原始股價
    elif (TARGET == Stock.Amazon): 
        # 將 'Open', 'High', 'Low', 'Close' 四个欄位的數據都乘以20
        selected_data[['Open', 'High', 'Low', 'Close']] = selected_data[['Open', 'High', 'Low', 'Close']] * 20

    # 計算 Return 新值：(開盤價 - 收盤價) / 開盤價，然後選取 'Open', 'High', 'Low', 'Volume' 欄位
    if (PRICE_Method == 'Return_Price'):
        selected_data['Close'] = (selected_data['Open'] - selected_data['Close']) / selected_data['Open']

    # 將'Close'以外的特徵都設為0，只用Close特徵訓練。
    columns_to_change = ['Open', 'High', 'Low', 'Volume']  # 這裡填入要更改的列名
    selected_data[columns_to_change] = 0

    # 顯示修改后的數據
    print(selected_data)

    # 儲存變數到Txt檔
    saveToTxt(selected_data,'selected_data.txt')

    # 繪製股價趨勢圖
    plot_stock_price_figures(data['Close'],'Stock Price Curve')
    # plot_close_price_trend(data)

    # ====建立模擬的新聞情感分數====
    days = len(selected_data)  # 模擬數據的天數
    # 生成模擬股價數據和新聞情感分數
    news_sentiment = np.random.uniform(-1, 1, days)  
    # 生成全為0的新聞情緒分數
    zero_sentiment = np.zeros(days)
    print (f"news_sentiment:{news_sentiment}")
    # 將news_sentiment轉換成二維數組
    simulation_sentiment = news_sentiment.reshape(-1, 1)  # 將 a 轉換為 (1413, 1) 的形狀
    zero_sentiment = zero_sentiment.reshape(-1, 1)  # 將 a 轉換為 (1413, 1) 的形狀
    print (f"news_sentiment reshape:{news_sentiment}")
    print (f"selected_data:{selected_data}")

    # 使用csv內Average Sentiment
    news_sentiment = data[['Average Sentiment']]
    # 合併數據與情感分數
    data_with_sentiment = np.column_stack((selected_data, news_sentiment))
    data_without_sentiment = np.column_stack((selected_data, zero_sentiment))

    # 儲存變數到Txt檔
    saveToTxt(news_sentiment,'data_news_sentiment.txt')
    saveToTxt(data_with_sentiment,'data_with_sentiment.txt')
    saveToTxt(data_without_sentiment,'data_without_sentiment.txt')

    #====正規化=======
    data_without_sentiment_normalized,scaler_without = normalize_data(data_without_sentiment,'without')
    data_with_sentiment_normalized,scaler_with = normalize_data(data_with_sentiment,'with')

    # 利用n天資料，預測第n+1天，每個序列的時間步長，n = time_step
    time_step = LSTM_Params['Time_Step']

    # [無情感分數]為不包含新聞情感分數的數據創建數據集
    X_without, Y_without = create_dataset(data_without_sentiment_normalized, time_step)
    # [有情感分數]為包含新聞情感分數的數據創建數據集
    X_with, Y_with = create_dataset(data_with_sentiment_normalized, time_step)

    # # 儲存變數到Txt檔，變數是3維資料，無法存成2維
    # saveToTxt(X_without,'X_without.txt')
    # saveToTxt(Y_without,'Y_without.txt')
    # saveToTxt(X_with,'X_with.txt')
    # saveToTxt(Y_with,'Y_with.txt')

    # [有問題]分割數據為訓練集和測試集，會對資料進行隨機洗牌，因為對於時間序列數據來說，保持時間順序是很重要的
    # X_train_without, X_test_without, Y_train_without, Y_test_without = train_test_split(X_without, Y_without, test_size=0.2, random_state=42)
    # X_train_with, X_test_with, Y_train_with, Y_test_with = train_test_split(X_with, Y_with, test_size=0.2, random_state=42)
    
    # 資料分割成80%訓練集、20%測試集，然後再從訓練集中切割出20%作為驗證集
    if (ModelType == Model.Training):
        # 定義訓練集和測試集的大小
        # 資料分割成80%訓練集、20%測試集，然後再從訓練集中切割出20%作為驗證集
        total_size = len(X_with)
        train_size = int(total_size * 0.8)
        test_size = total_size - train_size

        # [無情感分數]分割80%訓練集、20%測試集
        X_original_train_without, X_test_without = X_without[:train_size], X_without[train_size:]
        Y_original_train_without, Y_test_without = Y_without[:train_size], Y_without[train_size:]
        # 再從訓練集中切割出20%作為驗證集
        val_size = int(train_size * 0.2)
        X_train_without, X_val_without = X_original_train_without[:-val_size], X_original_train_without[-val_size:]
        Y_train_without, Y_val_without = Y_original_train_without[:-val_size], Y_original_train_without[-val_size:]

        # [有情感分數]分割80%訓練集、20%測試集
        X_original_train_with, X_test_with = X_with[:train_size], X_with[train_size:]
        Y_original_train_with, Y_test_with = Y_with[:train_size], Y_with[train_size:]
        # 再從訓練集中切割出20%作為驗證集
        val_size = int(train_size * 0.2)
        X_train_with, X_val_with = X_original_train_with[:-val_size], X_original_train_with[-val_size:]
        Y_train_with, Y_val_with = Y_original_train_with[:-val_size], Y_original_train_with[-val_size:]
        
        # 建立和訓練模型
        # 使用六個特徵訓練，1.'Open', 2.'High', 3.'Low', 4.'Close', 5.'Volume' 6. 'Sentiment'
        model_without, history_without = build_and_train_model(X_train_without, Y_train_without, X_val_without, Y_val_without, 6, time_step,"without")  # 不包含新聞情感分數0
        model_with, history_with = build_and_train_model(X_train_with, Y_train_with, X_val_with, Y_val_with, 6, time_step,"with")   # 特徵數更改為6，包含新聞情感分數

        # 繪製訓練和測試的損失曲線
        plot_loss_curves(history_without, 'Training and Validation Loss - Without Sentiment')
        plot_loss_curves(history_with, 'Training and Validation Loss - With Sentiment')

        # [Paper1]拿測試資料進行預測
        Y_pred_without = model_without.predict(X_test_without)
        Y_pred_with = model_with.predict(X_test_with)
        
        # # [Paper2]拿全部原始股價資料進行預測
        # Y_pred_without_all = model_without.predict(X_without)
        # Y_pred_with_all = model_with.predict(X_with)
        
        # # 使用模型預測訓練集和測試集
        # Y_pred_train = model_with.predict(X_train_with)  # 預測訓練集
        # Y_pred_test = model_with.predict(X_test_with)  # 預測測試集

    elif (ModelType == Model.Predict):
       
        # 加載模型
        model_with = load_model(model_file_with)
        model_without = load_model(model_file_without)
        # 指定測試預測資料
        X_test_without = X_without
        Y_test_without = Y_without
        X_test_with = X_with
        Y_test_with = Y_with
        # 使用模型進行預測
        Y_pred_without = model_without.predict(X_test_without)
        Y_pred_with = model_with.predict(X_test_with)
        # 儲存變數到Txt檔
        saveToTxt(Y_pred_without,'Y_pred_without.txt')
        saveToTxt(Y_pred_with,'Y_pred_with.txt')

    # =======[Paper1] 將預測值和真實值轉換回原始範圍=========
    # 注意：為了逆轉換，我們需要創建一個與預測值相同形狀的數組，但所有值都是0，除了倒數第三列是預測值
    # 在這段代碼中，我們首先對模型的預測結果進行轉換，使其形狀與進行 inverse_transform 所需的形狀一致。
    # 然後，我們使用 inverse_transform 方法將預測值和真實值轉換回原始範圍。
    # 這裡，我們將預測結果或真實值附加到一個全為零的數組上，這個數組的大小和我們最初用於正規化的數組相同，
    # 這是因為 inverse_transform 需要與 fit_transform 期間相同形狀的輸入。
    # 由於我們只關心 'Close' 價格（在這個例子中是最後第三列），所以要將Y_pred_without，插入最後第三列

    # 設定列的索引為倒數第三列
    close_column_index = -3

    # 將預測值和真實值進行逆歸一化並儲存到文件
    Y_pred_without_final = inverse_transform_and_save(
        scaler_without, Y_pred_without, data_without_sentiment_normalized.shape, close_column_index, 'Y_pred_without_final.txt'
    )

    Y_pred_with_final = inverse_transform_and_save(
        scaler_with, Y_pred_with, data_with_sentiment_normalized.shape, close_column_index, 'Y_pred_with_final.txt'
    )

    Y_test_without_final = inverse_transform_and_save(
        scaler_without, Y_test_without, data_without_sentiment_normalized.shape, close_column_index, 'Y_test_without_final.txt'
    )

    Y_test_with_final = inverse_transform_and_save(
        scaler_with, Y_test_with, data_with_sentiment_normalized.shape, close_column_index, 'Y_test_with_final.txt'
    )

    print ("======================")
    # =======將預測值和真實值轉換回原始範圍End==================

    # MSE和MAE存檔名稱
    results_file = os.path.join(save_path, '5.MSE_MAE_Result.txt')

    # # [有情感分數]計算MSE和MAE、MAPE並保存到同一文本文件
    calculated_mse_with, calculated_mae_with , mape_with, rmse_with = calculate_and_print_metrics(Y_test_with.ravel(), Y_pred_with.ravel(), 'Test With Sentiment', results_file)

    # [無情感分數]計算MSE和MAE、MAPE並保存到同一文本文件
    calculated_mse_without, calculated_mae_without, mape_without, rmse_without = calculate_and_print_metrics(Y_test_without.ravel(), Y_pred_without.ravel(), 'Test Without Sentiment', results_file)

    # # [有情感分數]計算MSE和MAE、MAPE並保存到同一文本文件
    original_calculated_mse_with, original_calculated_mae_with, original_mape_with, original_rmse_with = calculate_and_print_metrics(Y_test_with_final.ravel(), Y_pred_with_final.ravel(), 'Original Test With Sentiment', results_file)

    # # [無情感分數]計算MSE和MAE、MAPE並保存到同一文本文件
    original_calculated_mse_without, original_calculated_mae_without, original_mape_without, original_rmse_without = calculate_and_print_metrics(Y_test_without_final.ravel(), Y_pred_without_final.ravel(), 'Original Test Without Sentiment', results_file)

    # 調用 plot_combined_figures 函數來繪製比較圖
    plot_combined_figures(
        Y_test_without_final,  # 原始的股價數據
        Y_pred_with_final,     # 包含情感分數的預測股價
        Y_pred_without_final,  # 不包含情感分數的預測股價
        'Comparison of Stock Price Predictions',  # 圖表標題
    )


    # 還原訓練和測試資料
    # restored_train_with = inverse_transform_and_save(
    #     scaler_with, X_train_with, data_with_sentiment_normalized.shape, -3, 'restored_train_with.txt'
    # )
    # restored_test_with = inverse_transform_and_save(
    #     scaler_with, X_test_with, data_with_sentiment_normalized.shape, -3, 'restored_test_with.txt'
    # )
    
    # =================[Paper2]=======================
    if (Stock_Period == '2020_AMZN_5min'):
        # 1.[無情感分數]
        # 首先對訓練和測試資料進行預測
        pred_train_without = model_with.predict(X_train_without)  # 對訓練資料的預測
        pred_val_without = model_with.predict(X_val_without)  # 對驗證資料的預測
        pred_test_without = model_with.predict(X_test_without)  # 對測試資料的預測

        # 還原預測的訓練和驗證和測試資料
        restored_pred_train_without = inverse_transform_and_save(
            scaler_without, pred_train_without, data_without_sentiment_normalized.shape, -3, 'restored_pred_train_without.txt'
        )
        restored_pred_val_without = inverse_transform_and_save(
            scaler_without, pred_val_without, data_without_sentiment_normalized.shape, -3, 'restored_pred_val_without.txt'
        )
        restored_pred_test_without = inverse_transform_and_save(
            scaler_without, pred_test_without, data_without_sentiment_normalized.shape, -3, 'restored_pred_test_without.txt'
        )
        # 使用train_test_split將資料切分成80%訓練集和20%測試集，為確保資料的時間順序，設定 shuffle=False，因為時間序列資料通常不需要打亂。
        train_close, test_close = train_test_split(selected_data['Close'], test_size=0.2, random_state=42, shuffle=False)

        # 將三個預測資料按順序串接在一起，轉換為 DataFrame 以便於串接
        df_pred_train_without = pd.DataFrame(restored_pred_train_without, columns=['Prediction'])
        df_pred_val_without = pd.DataFrame(restored_pred_val_without, columns=['Prediction'])
        df_pred_test_without = pd.DataFrame(restored_pred_test_without, columns=['Prediction'])

        # 按順序串接
        df_combined_predictions_without = pd.concat([df_pred_train_without, df_pred_val_without, df_pred_test_without], ignore_index=True)

        # 使用新的函數繪製圖形
        plot_restored_predictions(
            scaler_without,
            train_close,
            test_close,
            df_combined_predictions_without['Prediction'],
            "AMAZON Stock Market Prices Forecast without Sentiment Signal"
        )

        # 2.[有情感分數]
        # 首先對訓練和測試資料進行預測
        pred_train_with = model_with.predict(X_train_with)  # 對訓練資料的預測
        pred_val_with = model_with.predict(X_val_with)  # 對驗證資料的預測
        pred_test_with = model_with.predict(X_test_with)  # 對測試資料的預測

        # 還原預測的訓練和驗證和測試資料
        restored_pred_train_with = inverse_transform_and_save(
            scaler_with, pred_train_with, data_with_sentiment_normalized.shape, -3, 'restored_pred_train_with.txt'
        )
        restored_pred_val_with = inverse_transform_and_save(
            scaler_with, pred_val_with, data_with_sentiment_normalized.shape, -3, 'restored_pred_val_with.txt'
        )
        restored_pred_test_with = inverse_transform_and_save(
            scaler_with, pred_test_with, data_with_sentiment_normalized.shape, -3, 'restored_pred_test_with.txt'
        )

        # 使用train_test_split將資料切分成80%訓練集和20%測試集，為確保資料的時間順序，設定 shuffle=False，因為時間序列資料通常不需要打亂。
        train_close, test_close = train_test_split(selected_data['Close'], test_size=0.2, random_state=42, shuffle=False)

        # 將三個預測資料按順序串接在一起
        # 轉換為 DataFrame 以便於串接
        df_pred_train_with = pd.DataFrame(restored_pred_train_with, columns=['Prediction'])
        df_pred_val_with = pd.DataFrame(restored_pred_val_with, columns=['Prediction'])
        df_pred_test_with = pd.DataFrame(restored_pred_test_with, columns=['Prediction'])

        # 按順序串接
        df_combined_predictions = pd.concat([df_pred_train_with, df_pred_val_with, df_pred_test_with], ignore_index=True)

        # 使用新的函數繪製圖形
        plot_restored_predictions(
            scaler_with,
            train_close,
            test_close,
            df_combined_predictions['Prediction'],
            "AMAZON Stock Market Prices Forecast with Sentiment Signal"
        )
    # =================[Paper2]=======================

    # LSTM參數和MSE、MAE結果
    results = {
        '日期': [formatted_now],  
        '股票': [TARGET.name],  # 股票名稱
        'Epochs': [LSTM_Params['Epochs']],
        'Batch_Size': [LSTM_Params['Batch_Size']],
        'Time_Step': [LSTM_Params['Time_Step']],

        '[原始]無情感分數MSE': [original_calculated_mse_without],  # 替換為實際計算結果
        '[原始]有情感分數MSE': [original_calculated_mse_with],  # 替換為實際計算結果
        '[原始]有情感分數MSE進步幅度' : [original_calculated_mse_without-original_calculated_mse_with],
        '[原始]無情感分數MAE': [original_calculated_mae_without],  # 替換為實際計算結果
        '[原始]有情感分數MAE': [original_calculated_mae_with],  # 替換為實際計算結果
        '[原始]有情感分數MAE進步幅度' : [original_calculated_mae_without-calculated_mae_with],
        '[原始]無情感分數MAPE(%)': [original_mape_without],  # 替換為實際計算結果
        '[原始]有情感分數MAPE(%)': [original_mape_with],  # 替換為實際計算結果

        '無情感分數MSE': [calculated_mse_without],  # 替換為實際計算結果
        '有情感分數MSE': [calculated_mse_with],  # 替換為實際計算結果
        '有情感分數MSE進步幅度' : [calculated_mse_without-calculated_mse_with],
        '無情感分數MAE': [calculated_mae_without],  # 替換為實際計算結果
        '有情感分數MAE': [calculated_mae_with],  # 替換為實際計算結果
        '有情感分數MAE進步幅度' : [calculated_mae_without-calculated_mae_with],
        '無情感分數MAPE(%)': [mape_without],  # 替換為實際計算結果
        '有情感分數MAPE(%)': [mape_with],  # 替換為實際計算結果

        '[原始]無情感分數RMSE': [original_rmse_without],  # 替換為實際計算結果
        '[原始]有情感分數RMSE': [original_rmse_with],  # 替換為實際計算結果
        '無情感分數RMSE': [rmse_without],  # 替換為實際計算結果
        '有情感分數RMSE': [rmse_with]  # 替換為實際計算結果

    }

    # 轉換結果為DataFrame
    new_data = pd.DataFrame(results)

    # 指定Excel文件路徑
    file_path = 'LSTM_結果.xlsx'

    # 檢查文件是否存在
    if os.path.exists(file_path):
        # 讀取現有數據
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            old_data = pd.read_excel(file_path)
            # 合併舊數據和新數據
            updated_data = pd.concat([old_data, new_data], ignore_index=True)
            # 保存更新後的數據回Excel文件
            updated_data.to_excel(writer, index=False, sheet_name='結果')
    else:
        # 文件不存在，直接寫入新數據
        new_data.to_excel(file_path, index=False, sheet_name='結果')

    # 使用openpyxl加載工作簿
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # 設置第一行的行高
    sheet.row_dimensions[1].height = 45  # 可以根據需要調整行高

    # 遍歷第一行的所有單元格，設置列寬和文字自動換行
    for col in sheet.iter_cols(min_row=1, max_row=1):
        for cell in col:
            # 設置列寬，这里以设置为20为例
            sheet.column_dimensions[cell.column_letter].width = 12
            # 設定單元格的對齊方式為水平居中和垂直居中
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # 設置文字自動換行
            cell.alignment = Alignment(wrapText=True)
        
    # 儲存更改
    workbook.save(file_path)

    print ('========================')
    print (f'完成!數據相關檔案儲存在./{file_path}')
    print (f'完成!相關圖檔檔案儲存在資料夾 {save_path}')

def plot_restored_predictions(scaler, original_train, original_test, pred, title):
    """
    繪製訓練資料、測試資料及其預測結果（還原為原始尺度）在同一個圖形中。
    """
    plt.figure(figsize=[15, 10])
    plt.plot(original_train, label='Training', color='green')
    plt.plot(original_test, label='Test', color='blue')
    plt.plot(pred, label='Prediction', color='red')
    plt.title(title)
    plt.xlabel('TimeStamps')
    plt.ylabel('Stock Price')
    plt.legend(loc='upper left')  # 將標籤放在左上角
    plt.grid(True, linestyle='-', color='gray', linewidth=0.5)
    file_name = f'{save_path}3.2 {title}.png'
    plt.savefig(file_name, format='png', dpi=300)
    plt.close()

def plot_combined_predictions(training_data, test_data, pred_train, pred_test, title):
    """
    繪製訓練資料、測試資料及其預測結果在同一個圖形中。
    :param training_data: 訓練資料
    :param test_data: 測試資料
    :param pred_train: 對訓練資料的預測結果
    :param pred_test: 對測試資料的預測結果
    :param title: 圖形的標題
    """
    plt.figure(figsize=[15, 10])
    plt.plot(training_data, label='Training Data', color='blue')  # 訓練資料
    plt.plot(test_data, label='Test Data', color='orange')  # 測試資料
    plt.plot(pred_train, label='Prediction on Training Data', color='green')  # 對訓練資料的預測
    plt.plot(pred_test, label='Prediction on Test Data', color='red')  # 對測試資料的預測
    plt.title(title)
    plt.xlabel('TimeStamps')
    plt.ylabel('Stock Price')
    plt.legend()
    plt.grid(True, linestyle='-', color='gray', linewidth=0.5)  # 加入灰色格線
    file_name = f'{save_path}combined_{title}.png'  # 設定儲存路徑
    plt.savefig(file_name, format='png', dpi=300)  # 儲存為高解析度PNG檔案
    plt.close()  # 關閉圖形

# 通用函數：將預測值或真實值轉換回原始範圍並保存到文件
def inverse_transform_and_save(scaler, predictions, data_normalized_shape, column_index, output_file):
    print("預測資料predictions的長度:", len(predictions))  # 檢查預測資料的長度
    print("預測資料predictions的形狀:", predictions.shape)  # 檢查預測資料的長度
    print("目標陣列data_normalized_shape的形狀:", data_normalized_shape)  # 確保目標陣列的形狀正確
    print("列索引:", column_index)  # 檢查列索引

    # 根據預測資料的長度創建新的陣列，確保長度匹配
    new_array = np.zeros((len(predictions), data_normalized_shape[1]))  # 確保陣列的形狀正確
    print("new_array的形狀:", new_array.shape)  # 確保目標陣列的形狀正確
    new_array[:, column_index] = predictions.ravel()  # 將預測結果放入指定列
    print("new_array的值:", new_array)
    # 進行反轉換
    original_values = scaler.inverse_transform(new_array)

    # 取得指定列的資料
    final_values = original_values[:, column_index]

    # 將結果儲存到指定的文件
    # np.savetxt(output_file, final_values, fmt='%f')
    saveToTxt(final_values,output_file)

    return final_values

def saveToTxt(data,fileName):
    '''儲存變數到Txt檔'''
    # 文件的完整路徑
    file_name = f'{save_path}/{fileName}'
    # 儲存文件
    np.savetxt(file_name, data, fmt='%f')  # 使用您的數據替換此處

def plot_close_price_trend(data, dpi=300):
    """
    從CSV檔案讀取收盤價格並繪製股價趨勢圖。
    參數:
    csv_file_path (str): CSV檔案的路徑。
    output_png_path (str): 圖形將要被保存的PNG檔案路徑。
    dpi (int): 輸出圖片的每英吋點數，用於高解析度。預設值為300。
    """
    # # 從CSV檔案中讀取數據
    # data = pd.read_csv(csv_file_path)
    # 確保日期格式正確
    # data['Date'] = pd.to_datetime(data['Date'])
    
    # 繪製收盤價趨勢圖
    plt.figure(figsize=[10, 5])
    plt.plot(data['Date'], data['Close'], label='Close Price', color='blue')
    plt.title('Close Price Trend')
    plt.xlabel('TimeStamps Days')
    plt.ylabel('Close Price')
    plt.xticks(rotation=45)  # 日期旋轉以改善可讀性
    
    data['Date'] = pd.to_datetime(data['Date'])
    # 將日期轉換為開始日期以來的天數
    data['Days'] = (data['Date'] - data['Date'].min()).dt.days

    # 設定X軸標籤間隔每150天顯示一次
    max_days = data['Days'].max()
    plt.xticks(np.arange(0, max_days+1, 150))  # 從0開始，到最大天數，每150天標一個點

    plt.legend()
    plt.grid(True)
    plt.tight_layout()  # 調整布局以確保標籤不被切割
    file_name = f'{save_path}/1.{TARGET}_close_price_trend.png'

    # 將圖形保存為高解析度的PNG檔案
    plt.savefig(file_name, format='png', dpi=dpi) # 將圖形保存為PNG檔案
    plt.close()  # 關閉圖形


# ================添加預測模式===================
def normalize_data(data,title):
    if (Stock_Period == '2020_AMZN_5min'):
          # 正規化資料-1~1之間
        scaler = MinMaxScaler(feature_range=(-1, 1))
    else:
        # 正規化資料0~1之間
        scaler = MinMaxScaler(feature_range=(0, 1))

    data_normalized = scaler.fit_transform(data)

    # 儲存變數到Txt檔
    saveToTxt(data,f'data_{title}.txt')
    saveToTxt(data_normalized,f'data_{title}_normalized.txt')

    return data_normalized, scaler

def convert_to_lstm_format(data, time_step=1):
    X = []
    for i in range(len(data) - time_step):
        a = data[i:(i + time_step), :]
        X.append(a)
    return np.array(X)

# =====轉換數據為LSTM預期的格式======
# 轉換數據為LSTM預期的格式
def create_dataset(data, time_step=1):
    X, Y = [], []
    for i in range(len(data) - time_step):
        a = data[i:(i + time_step), :]  # 包括最後一列的特徵
        # 文件的完整路徑
        file_name = f'{save_path}/create_dataset_X.txt'
        # 使用savetxt函數儲存數組
        np.savetxt(file_name, a, fmt='%f')  # 這裡的fmt參數指定了格式，'%d'意味著整數格式
        b = data[i + time_step, -3]
        # 確保 b 是一維數組
        b_array = np.array([b])  # 將 b 轉換為一維數組
        # 文件的完整路徑
        file_name = f'{save_path}/create_dataset_Y.txt'
        np.savetxt(file_name, b_array, fmt='%f')  # 這裡的fmt參數指定了格式，'%d'意味著整數格式
        
        X.append(a)
        Y.append(data[i + time_step, -3])  # 取 'Close' 為預測目標
    return np.array(X), np.array(Y)


# 定義一個函數來構建和訓練LSTM模型
def build_and_train_model(X_train, Y_train, X_test, Y_test, features, time_step, name):
    # 重塑輸入為[LSTM的樣本數, 時間步長, 特徵數]
    X_train = np.reshape(X_train, (X_train.shape[0], time_step, features))
    X_test = np.reshape(X_test, (X_test.shape[0], time_step, features))

    # 建立LSTM模型
    model = Sequential([
        LSTM(50, return_sequences=True, input_shape=(time_step, features)),
        LSTM(50),
        Dense(1)
    ])
    model.compile(loss='mean_squared_error', optimizer='adam')

    model.summary()
    
    # 訓練模型
    history = model.fit(X_train, Y_train, epochs=LSTM_Params['Epochs'], batch_size=LSTM_Params['Batch_Size'], validation_data=(X_test, Y_test), verbose=2)
    # 保存模型到指定路徑
    model.save(f'{save_path}6.{formatted_now}_{TARGET}_{PRICE_Method}_{name}_model.h5')  
    return model, history

# 定義一個函數來繪製訓練和測試的損失曲線
def plot_loss_curves(history, title):
    plt.figure()
    plt.plot(history.history['loss'], label='Training Loss')
    plt.plot(history.history['val_loss'], label='Validation Loss')
    plt.title(title)
    plt.xlabel('Epochs')
    plt.ylabel('Loss')
    plt.legend()

    # 將Training Loss和Validation Loss保存到同一個TXT文件中
    with open(f'{save_path}/4.{title}_loss.txt', 'w') as f:
        f.write("Epoch,Training Loss,Validation Loss\n")  # 添加標題行
        for i in range(len(history.history['loss'])):
            f.write(f"{i+1},{history.history['loss'][i]},{history.history['val_loss'][i]}\n")

    # # 保存損失數據到TXT文件
    # with open(f'{save_path}/4.{title}.txt', 'w') as f:
    #     for loss in history.history['loss']:
    #         f.write(f"{loss}\n")

    # with open(f'{save_path}/4.{title}.txt', 'w') as f:
    #     for loss in history.history['val_loss']:
    #         f.write(f"{loss}\n")

    # 保存圖表
    file_name = f'{save_path}2.{title}.png'
    plt.savefig(file_name, format='png', dpi=300)  # 儲存高精度PNG檔案
    # plt.show()
    plt.close()  # 關閉繪圖視窗

def calculate_mape(true_values, predicted_values):
    # MAPE計算，避免除以零
    # 使用tf.clip_by_value來避免除以零的情況，這是計算MAPE時常見的問題。
    # 我們將true_values和predicted_values轉換為浮點數，以確保計算過程中數據類型的一致性。
    true_values = tf.cast(true_values, tf.float32)
    predicted_values = tf.cast(predicted_values, tf.float32)
    # 使用 tf.clip_by_value 來避免除以一個非常小的數
    # 注意，我們把下限設為 1e-2，這可以根據實際情況調整
    diff = tf.abs((true_values - predicted_values) / tf.clip_by_value(true_values, 1e-1, tf.float32.max))
    # 有問題，預測時mape算出來的值會超過100%
    # diff = tf.abs((true_values - predicted_values) / tf.clip_by_value(true_values, 1e-8, tf.float32.max))
    mape = tf.reduce_mean(diff) * 100
    return mape.numpy()

def calculate_and_print_metrics(true_values, predicted_values, title, file_name):
    # 定義一個函數來計算和打印MSE和MAE，並將結果保存到相同的文本文件中
    mse = tf.keras.losses.MeanSquaredError()
    mae = tf.keras.losses.MeanAbsoluteError()
    calculated_mse = mse(true_values, predicted_values).numpy()
    calculated_mae = mae(true_values, predicted_values).numpy()
    calculated_mape = calculate_mape(true_values, predicted_values)
    calculated_rmse = np.sqrt(calculated_mse)  # 計算RMSE

    result_string = f'{title} - MSE: {calculated_mse}, RMSE:{calculated_rmse} ,MAE: {calculated_mae}, MAPE: {calculated_mape}%\n'
    
     # 將結果保存到文本文件
    with open(file_name, 'a') as file:  # 'a'模式表示附加內容
        file.write(result_string)

    print(f'{title} - MSE: {calculated_mse}, RMSE:{calculated_rmse}, MAE: {calculated_mae}, MAPE: {calculated_mape}%')
    return calculated_mse, calculated_mae, calculated_mape, calculated_rmse

def plot_stock_price_figures(data, title):
    plt.figure(figsize=[15,10])
    plt.plot(data, label='Original Stock Price', color='blue')
    plt.title(title)
    plt.xlabel('TimeStamps')
    plt.ylabel('Stock Price')
    plt.legend()
    plt.grid(True, linestyle='-', color='gray', linewidth=0.5)  # 加入灰色格線
    file_name = f'{save_path}1.{title}.png'
    plt.savefig(file_name, format='png', dpi=300)  # 儲存高精度PNG檔案
    # plt.show()  # 顯示圖像

def plot_combined_figures(true_values, pred_with_sentiment, pred_without_sentiment, title):
    ''' 畫出[Paper1] 1.原始股價 2.預測Sentiment結果 3.預測without Sentiment結果'''
    plt.figure(figsize=[15,10])
    plt.plot(true_values, label='Original Stock Price', color='blue')
    plt.plot(pred_with_sentiment, label='Predicted with Sentiment', color='orange')
    plt.plot(pred_without_sentiment, label='Predicted without Sentiment', color='green')
    plt.title(title)
    plt.xlabel('TimeStamps')
    plt.ylabel('Stock Price')
    plt.legend()
    plt.grid(True, linestyle='-', color='gray', linewidth=0.5)  # 加入灰色格線
    file_name = f'{save_path}3.1{title}.png'
    plt.savefig(file_name, format='png', dpi=300)  # 儲存高精度PNG檔案
    # plt.show()  # 顯示圖像

if __name__ == '__main__':
    MAIN(file_path)